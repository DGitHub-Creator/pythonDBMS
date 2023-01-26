import os

from openpyxl import *

if not os.path.exists('data/table_authority.xlsx'):  # 创建表的权限表
    Workbook().save('data/table_authority.xlsx')

dbname = 'system'
tbauth = load_workbook('data/table_authority.xlsx')  # 加载数据表的权限表
tbauth_tb = tbauth.create_sheet(dbname)
authority_tbcol = ['database', 'table', 'select', 'insert', 'delete', 'update', 'use']  # 初始化权限项目
for i in range(len(authority_tbcol)):  # 创建表头
    tbauth_tb.cell(row=1, column=i + 1).value = authority_tbcol[i]

if tbauth.worksheets[0].title == 'Sheet':
    del tbauth['Sheet']
tbauth.save('data/table_authority.xlsx')

tbauth = load_workbook('data/table_authority.xlsx')
if dbname + '1' in tbauth.sheetnames:
    del tbauth[dbname]
    tbauth[dbname + '1'].title = dbname
tbauth.save('data/table_authority.xlsx')

print('数据库 ' + dbname + ' 创建操作执行成功。')
