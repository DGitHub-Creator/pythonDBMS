import re

import pandas as pd
from openpyxl import *


def grant(using_dbname, auths, type_au, names, users, options):
    try:
        dff = pd.read_excel('data/system.xlsx')
        all_users = dff['username'].values.tolist()
    except:
        return
    print(users)
    print(all_users)
    user_exist = True
    for user in users:
        if user not in all_users:
            user_exist = False
            print('[!]用户', user, "不存在！")
    if user_exist is False:
        print('[!]请检查输入的用户名是否存在！')
        return
    try:
        if auths == 'ALL PRIVILEGES':  # 所有权限
            auths = ['create', 'select', 'insert', 'delete', 'update', 'use', 'grant', 'revoke']
    except:
        ''
    try:
        if users.lower() == 'public':  # 公共用户
            df = pd.read_excel('data/system.xlsx')
            users = df['username'].values.tolist()
    except:
        ''
    if type_au.upper() == 'TABLE':
        audb = load_workbook('data/table_authority.xlsx')  # 加载数据表的权限表
        try:
            autab = audb[using_dbname]  # 打开对应数据库的数据表的权限表
            # print(autab)
        except:
            print('[!]没有', using_dbname, '数据库')
            return
        aumax_row = autab.max_row  # 获取最大行
        aumax_col = autab.max_column  # 获取最大列
        for table_name in names:  # 循环找到行号
            rw = -1  # 标记行
            for i in range(2, aumax_row + 1):
                if autab.cell(row=i, column=2).value == table_name:  # 找到表名对应的行数
                    rw = i
                    if autab.cell(row=rw, column=1).value != using_dbname:  # 如果不是本数据库的同名数据就跳出循环找到下一个
                        continue
                    else:  # 如果是本数据库的数据表就打断循环，即找到了表名对应的行数
                        break
            for auth in auths:  # 权限有多个，开始循环选择权限
                cl = -1  # 标记列
                for i in range(1, aumax_col + 1):
                    if autab.cell(row=1, column=i).value == auth.lower():  # 找到权限对应的列数
                        cl = i  # 找到了
                        break
                if rw == -1:
                    print('[!]没有找到此数据表！')
                    return
                if cl == -1:
                    print('[!]没有找到此权限！')
                    return
                old_user = autab.cell(row=rw, column=cl).value  # 获取原来的用户名
                # 开始把用户名拼接到旧的用户名上
                if old_user is None:  # 如果没有被授权过
                    new_user = users[0]
                else:  # 如果授权过
                    new_user = old_user + ',' + users[0]
                for i in range(1, len(users)):  # 把用户名拼接到旧的用户名上
                    new_user = new_user + ',' + users[i]
                # print(new_user)
                autab.cell(row=rw, column=cl).value = new_user  # 回填
        audb.save('data/table_authority.xlsx')  # 保存
        print('[*]授权成功！')
    elif type_au.upper() == 'DATABASE':
        audb = load_workbook('data/system.xlsx')  # 加载数据表的权限表
        try:
            autab = audb['authority']  # 打开对应数据库的数据表的权限表
            # print(autab)
        except:
            print('[!]没有', 'authority', '数据表')
            return
        aumax_row = autab.max_row  # 获取最大行
        aumax_col = autab.max_column  # 获取最大列
        for db_name in names:  # 循环找到行号
            rw = -1  # 标记行
            for i in range(2, aumax_row + 1):
                if autab.cell(row=i, column=1).value == db_name:  # 找到表名对应的行数
                    rw = i
                    break
            for auth in auths:  # 权限有多个，开始循环选择权限
                cl = -1  # 标记列
                for i in range(1, aumax_col + 1):
                    if autab.cell(row=1, column=i).value == auth.lower():  # 找到权限对应的列数
                        cl = i  # 找到了
                        break
                if rw == -1:
                    print('[!]没有找到此数据库！')
                    return
                if cl == -1:
                    print('[!]没有找到此权限！')
                    return
                old_user = autab.cell(row=rw, column=cl).value  # 获取原来的用户名
                # 开始把用户名拼接到旧的用户名上
                if old_user is None:  # 如果没有被授权过
                    new_user = users[0]
                else:  # 如果授权过
                    new_user = old_user + ',' + users[0]
                for i in range(1, len(users)):  # 把用户名拼接到旧的用户名上
                    new_user = new_user + ',' + users[i]
                # print(new_user)
                autab.cell(row=rw, column=cl).value = new_user  # 回填
        audb.save('data/system.xlsx')  # 保存
        print('[*]授权成功！')
        pass
    else:
        print('[!]授权错误！请检查输入后重试！')
        return


def sql_Analysis(using_dbname, sql, tag=''):
    sql = sql.strip()
    sql_word = sql.split(' ')  # 分割sql语句
    if len(sql_word) < 2:
        print('[!]语法错误')
        return
    # print(sql_word)
    operate = sql_word[0].lower()  # 操作名
    # GRANT SELECT ON TABLE Student TO U1
    # GRANT ALL PRIVILEGES ON TABLE Student,Course TO U2,U3
    # GRANT SELECT ON TABLE Student,Course TO PUBLIC

    # GRANT UPDATE(Sno),SELECT ON TABLE Student TO U4
    # GRANT INSERT ON TABLE Student TO U5 WITH GRANT OPTION
    if operate == 'grant':  # 如果是授权
        # print(operate.upper())
        auths = re.findall('GRANT (.*) ON', sql.upper())[0]  # 提取权限
        # print(auths)
        if auths == 'ALL PRIVILEGES':  #
            type_au = sql_word[4]
            names = sql_word[5].split(',')  # 获取表名
            users = sql_word[7].split(',')  # 获取用户名
            try:
                options = sql_word[8:]  # 获取传递权限
            except:
                options = ' '
        else:
            auths = auths.split(',')  # 提取权限
            type_au = sql_word[3]
            names = sql_word[4].split(',')  # 获取表名
            users = sql_word[6].split(',')  # 获取用户名
            try:
                options = sql_word[7:]  # 获取传递权限
            except:
                options = ' '
        # print(auths)
        # print(type_au)
        # print(tables)
        # print(users)
        # print(options)
        grant(using_dbname, auths, type_au, names, users, options)  # 授权


if __name__ == '__main__':
    for i in range(10):
        using_dbname = 'S-T'
        # using_db = pd.read_excel('data/' + using_dbname + '.xlsx', sheet_name=None)
        sql = input('[!][' + str(i) + ']>> ')
        sql_Analysis(using_dbname, sql)
