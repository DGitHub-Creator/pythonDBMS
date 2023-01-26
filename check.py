import pandas as pd
from openpyxl import *


def check_unique(dic, table, table_name, using_dbname):
    """
    检查是否唯一
    :param dic: 字典类型：{列名，数据名}
    :param table:数据表
    :param table_name:数据表的名称
    :param using_dbname: 数据库名称
    :return:
    """
    db_info = pd.read_excel('data/table_information.xlsx', sheet_name=None)
    table_info = db_info[using_dbname]
    tablene = table_info[table_info['table'] == table_name]
    tablepk = tablene[(tablene['primary_key']) == '1']
    if tablepk.empty:
        tablepk = tablene[(tablene['primary_key']) == 1]
    # print(tablepk)
    try:
        uqname_pk = tablepk['column_name'].values.tolist()
    except:
        print(table_name, '表中没有查询到主码，请检查主码设置')
        return False
    print(table_name, '表的主码(pk)为：', uqname_pk)
    tableuq = tablene[(tablene['unique']) == '1']
    if tableuq.empty:
        tableuq = tablene[(tablene['unique']) == 1]
    # print(tableuq)
    try:
        uqname_fg = tableuq['column_name'].values.tolist()
    except:
        uqname_fg = []
    print(table_name, '表中的unique值为：', uqname_fg)
    uqname_old = uqname_pk + uqname_fg
    uqname = list(dict.fromkeys(uqname_old))
    print(':::', uqname)
    print('---------------------')
    #     print(table)
    for key in uqname:
        count_tablevalue = table[key].values.tolist()
        # print(count_tablevalue)
        try:
            if int(dic[key]) in count_tablevalue:
                print(dic[key], '重复！')
                return False
        except:
            if dic[key] in count_tablevalue:
                print(dic[key], '重复！')
                return False
    print('是否唯一检查约束成功！')  # 显示 检查完成
    return True


def check_Constraint(using_dbname, dic_type, table_name):
    """
    插入时检查实体完整性约束，包括主码不为空，主码唯一，各数据对应的类型一致
    :param using_dbname: 数据库名称
    :param dic_type: 字典类型，{列名：数据类型}
    :param table_name:数据表的名称
    :return:
    """
    db_info = pd.read_excel('data/table_information.xlsx', sheet_name=None)  # 读入数据表的信息
    table_info = db_info[using_dbname]  # 查找到对应的数据库的信息表
    tablene = table_info[table_info['table'] == table_name]  # 查找到对应的数据表的信息
    tablepk = tablene[(tablene['primary_key']) == '1']  # 查找主码，主码为primary_key标记为1，这个1可能是字符类型也可能是数字类型
    if tablepk.empty:  # 如果没有查找到字符类型的标记，就查找数字类型的标记
        tablepk = tablene[(tablene['primary_key']) == 1]
    # print(tablepk)  # 显示 主码，可能为空
    try:
        pkname = tablepk['column_name'].tolist()  # 获取主码
        # print(pkname)  # 显示主码
    except:
        print(table_name, '表中没有查询到主码，请检查主码设置')  # 报错 没有主码
        return False
    print(table_name, '表的主码(pk)为：', pkname)
    try:
        for pk in pkname:
            if dic_type[pk] is None:  # 要插入的值中主码为空
                print('[!]语法错误!：主码为空！')
                return False
    except:
        print('[!]语法错误!：缺失主码！')  # 要插入的值中缺失主码
        return False
    print('---------------------')  # 显示 分隔符
    # print(tablene)
    # 将信息表里面的列名和类型打包成字典
    ne_name = tablene['column_name'].values.tolist()  # 列名转化为列表
    # print(ne_name)
    ne_type = tablene['type'].values.tolist()  # 类型转化为列表
    # print(ne_type)
    ne_dic = dict(map(lambda x, y: [x, y], ne_name, ne_type))  # 按对应关系打包
    # print(ne_dic)
    column_dic = dic_type.keys()  # 获取要插入数据的列名
    # print(column_dic)
    for key in column_dic:
        # print(dic_type[key],'==',ne_dic[key])
        if dic_type[key] != ne_dic[key]:  # 检查要插入的类型与信息表里面的类型是够对称
            print('[!]数据类型错误!：', key, '的数据类型是', ne_dic[key], '而不是', dic_type[key])
            return False
    print('实体完整性检查约束成功！')  # 显示 检查完成
    return True


# def check_tableAuth(user, dbname, tablename, operate):
#     '''
#     检查数据表的权限
#     :param user:待检查的用户
#     :param tablename: 用户需要访问的数据表
#     :param operate: 用户申请的操作
#     :return:
#     '''
#     users_row = -1
#     users_col = -1
#     db = load_workbook('data/table_authority.xlsx')  # 读取系统数据库里面的权限表
#     table = db[dbname]
#     max_row = table.max_row
#     max_col = table.max_column
#     # 查找对应的数据表和对应的操作的位置(对应的位置存放着允许执行这个表中的这个操作的用户名)
#     for i in range(1, max_row + 1):
#         if table.cell(row=i, column=2).value == tablename:
#             users_row = i
#             # print(users_row)
#             break
#     for j in range(1, max_col + 1):
#         if table.cell(row=1, column=j).value == operate:
#             users_col = j
#             # print(users_col)
#             break
#     # 读取允许执行这个表中的这个操作的用户名
#     try:
#         allow_users = table.cell(row=users_row, column=users_col).value.split(',')
#     except IOError:
#         return
#     if user in allow_users:  # 如果列表中存在，则允许此操作
#         print("您有权限访问 {} 数据表!".format(tablename))
#         return True
#     else:  # 如果列表中不存在，则此操作不被允许
#         print("您没有权限访问 {} 数据表！请检查后重试。".format(tablename))
#         return False
#     pass


def check_table_Authority(user, dbname, tablename, operate):
    """
    检查权限
    :param user:待检查的用户
    :param dbname: 用户需要访问的数据库
    :param tablename: 用户需要访问的数据表
    :param operate: 用户申请的操作
    :return:
    """
    db = load_workbook('data/table_authority.xlsx')  # 读取系统数据库里面的权限表
    table = db[dbname]
    max_row = table.max_row
    max_col = table.max_column
    # 查找对应的数据表和对应的操作的位置(对应的位置存放着允许执行这个表中的这个操作的用户名)
    for i in range(1, max_row + 1):
        if table.cell(row=i, column=2).value == tablename:
            users_row = i
            # print(users_row)
            break
    for j in range(2, max_col + 1):
        if table.cell(row=1, column=j).value == operate:
            users_col = j
            # print(users_col)
            break
    # 读取允许执行这个表中的这个操作的用户名
    allow_users = table.cell(row=users_row, column=users_col).value.split(',')
    if user in allow_users:  # 如果列表中存在，则允许此操作
        print("您有 {} 权限操作 {} 数据库的 {} 数据表!".format(operate, dbname, tablename))
        return True
    else:  # 如果列表中不存在，则此操作不被允许
        print("[!]您没有 {} 权限操作 {} 数据库的 {} 数据表!请检查后重试。".format(operate, dbname, tablename))
        return False
    pass


if __name__ == '__main__':
    user = 'admin111'
    dbname = 'S-T'
    tablename = 'Student'
    operate = 'select'
    check_table_Authority(user, dbname, tablename, operate)
