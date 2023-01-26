import hashlib
import os
import re

import pandas as pd
from openpyxl import *

import delete as dl
import grant as gt
import help as hl
import insert as ins
import revoke as rv
import select as sl
import update as ud

using_dbname = ''
db_path = 'data/'


def welcome():
    '''
    显示欢迎界面
    :return:
    '''
    print("""
--------------------------------------------------                                                                                                                  
 _       __       __                            
| |     / /___   / /_____ ____   ____ ___   ___ 
| | /| / // _ \ / // ___// __ \ / __ `__ \ / _ \\
| |/ |/ //  __// // /__ / /_/ // / / / / //  __/
|__/|__/ \___//_/ \___/ \____//_/ /_/ /_/ \___/ 
              -> 请先登录您的账号 <-
          -> help:语法帮助  exit:退出 <-
--------------------------------------------------     
          """)


def Init():
    '''
    初始化，检测数据目录和文件是否存在，没有存在就创建目录和文件
    :return:
    '''
    if not os.path.exists(db_path):
        os.mkdir(db_path)
    if not os.path.exists('data/table_information.xlsx'):  # 创立表的信息表
        Workbook().save('data/table_information.xlsx')
    if not os.path.exists('data/table_authority.xlsx'):  # 创建表的权限表
        Workbook().save('data/table_authority.xlsx')
    if os.path.exists('data/table_information.xlsx'):  # 删除残余的表项
        db = load_workbook('data/table_information.xlsx')
        sheet_name = db.sheetnames
        for name in sheet_name:
            if not os.path.exists('data/' + name + '.xlsx'):  # 如果不存在这个表，就把这个表的依赖删除掉
                del db[name]
        # db.save('data/table_information.xlsx')
    if os.path.exists('data/table_authority.xlsx'):  # 删除残余的表项
        db = load_workbook('data/table_authority.xlsx')
        sheet_name = db.sheetnames
        for name in sheet_name:
            if not os.path.exists('data/' + name + '.xlsx'):  # 如果不存在这个表，就把这个表的依赖删除掉
                del db[name]
    if os.path.exists('data/system.xlsx'):  # 删除残余的表项
        db = load_workbook('data/system.xlsx')
        sheet_name = db.sheetnames
        for name in sheet_name:
            if not os.path.exists('data/' + name + '.xlsx'):  # 如果不存在这个表，就把这个表的依赖删除掉
                del db[name]
        # db.save('data/system.xlsx')
        print('初始化中……………………')  # 初始化信息
    else:
        creat_db('system')  # 创建系统表，其中有用户名和数据库的权限
    db = load_workbook('data/system.xlsx')  # 读取系统表
    authority_tbcol = ['database char[50] pk unique', 'create char', 'select char', 'insert char', 'delete char',
                       'update char',
                       'use char', 'grant char', 'revoke char']  # 初始化权限项目
    creat_table('authority', db, 'system', authority_tbcol)  # 创建权限表，并初始化系统权限


def create_tbinfo(dbname):
    '''
    创建数据库信息存放表
    :param dbname: 数据库名称
    :return:
    '''
    db = load_workbook('data/table_information.xlsx')  # 读取信息存放数据库
    table = db.create_sheet(dbname)  # 创建对应数据库的信息存放表
    columns_name = ['table', 'column_name', 'type', 'null', 'unique', 'primary_key', 'foreign_key']  # 数据库的基本信息
    for i in range(len(columns_name)):  # 挨个写入表头
        table.cell(row=1, column=i + 1).value = columns_name[i]
    if db.worksheets[0].title == 'Sheet':  # 删除新建excel的默认sheet表
        del db['Sheet']
    db.save("data/table_information.xlsx")  # 保存信息
    print(dbname, 'tbinfo-初始化成功。')  # 显示 数据库信息初始化成功


def creat_db(dbname):
    '''
    创建数据库DATABASE，且初始化admin账号
    :param dbname: 数据库名称
    :return:
    '''
    dbpath = 'data/' + dbname + '.xlsx'  # 数据库的路径
    database = Workbook()  # 创建一个空的数据库
    if dbname == 'system':  # 如果是系统数据库
        # 初始化admin账号
        database.create_sheet('user')  # 创建用户表
        table = database['user']  # 读取用户表
        table.cell(row=1, column=1).value = 'username'  # 用户名列名
        table.cell(row=1, column=2).value = 'password'  # 密码列名
        table.cell(row=2, column=1).value = 'admin'  # 写入管理员账号
        table.cell(row=2, column=2).value = hashlib.md5('admin'.encode('utf-8')).hexdigest()  # 写入管理员密码并用MD5加密
    database.save(dbpath)  # 保存数据库
    create_tbinfo(dbname)  # 创建数据库的信息
    # 初始化数据库的权限
    if dbname != 'system':  # 创建system数据库时不需要创建时授予权限，而在创建authority表时会初始化system表和table_information表权限的
        if user not in ['admin', 'root']:  # 授予权限
            auth_name = 'admin,root' + ',' + user  # 追加当前用户的权限，即此数据库的创建者
        else:
            auth_name = 'admin,root'
        db = load_workbook('data/system.xlsx')  # 打开系统数据库，里面有数据库权限表，给与创建者使用权限
        sheet_name = db.sheetnames
        # if 'authority' not in sheet_name:
        #     db.create_sheet('authority')
        authority = db['authority']  # 权限表在创建system表时已经创建
        max_row = authority.max_row  # 获取最大行
        max_col = authority.max_column
        authority.cell(row=max_row + 1, column=1).value = dbname  # 写入数据库名称
        for i in range(2, max_col + 1):
            authority.cell(row=max_row + 1, column=i).value = auth_name  # 授予此用户权限
        db.save('data/system.xlsx')  # 保存
    tbauth = load_workbook('data/table_authority.xlsx')  # 加载数据表的权限表
    tbauth_tb = tbauth.create_sheet(dbname)
    authority_tbcol = ['database', 'table', 'create', 'select', 'insert', 'delete', 'update', 'use', 'grant',
                       'revoke']  # 初始化权限项目
    for i in range(len(authority_tbcol)):  # 挨个写入表头
        tbauth_tb.cell(row=1, column=i + 1).value = authority_tbcol[i]
    if dbname == 'system':
        tbauth_tb.cell(row=2, column=1).value = 'system'
        tbauth_tb.cell(row=2, column=2).value = 'user'
        max_row = tbauth_tb.max_row
        max_col = tbauth_tb.max_column
        for row in range(2, max_row + 1):
            for col in range(3, max_col + 1):
                tbauth_tb.cell(row=row, column=col).value = 'admin,root'
    if tbauth.worksheets[0].title == 'Sheet':  # 删除创建excel的默认sheet表
        del tbauth['Sheet']
    tbauth.save('data/table_authority.xlsx')  # 保存一下
    tbauth = load_workbook('data/table_authority.xlsx')  # 如果有重名的，就要删除掉重名的旧表
    if dbname + '1' in tbauth.sheetnames:  # 如果表里面有重名的旧表，创建新表时名称会自动加‘1’，比如dbname1，所以要删除掉1
        del tbauth[dbname]  # 删除掉旧表
        tbauth[dbname + '1'].title = dbname  # 更改名称
    tbauth.save('data/table_authority.xlsx')  # 保存
    print('数据库 ' + dbname + ' 创建操作执行成功。')


def creat_table(table_name, current_db, current_dbname, column_list):
    '''
    创建数据表
    :param table_name: 数据表名称
    :param current_db: 数据库
    :param current_dbname: 数据库名称
    :param column_list:数据表的列名称(表头)
    :return:
    '''
    # print('current_db.sheetnames:', current_db.sheetnames)
    if table_name not in current_db.sheetnames:  # 创建数据表
        table = current_db.create_sheet(table_name)
    else:
        if current_dbname == 'system':
            print('system 数据表 已存在')
            return
        else:
            print('数据表 已存在，请重新输入。')
            return
    if current_db.worksheets[0].title == 'Sheet':  # 删除掉创建excel时默认的sheet表
        del current_db['Sheet']
    length = len(column_list)  # 列（表头）的数量
    # 将表头的信息转化存放到信息表table_information
    tbinfo = load_workbook('data/table_information.xlsx')  # 打开信息数据库
    tbinfo_tb = tbinfo[current_dbname]  # 打开对应的数据表
    tbinfo_rows = tbinfo_tb.max_row  # 获取最大行数
    column_names = []  # 获取表头信息中的列名
    # 循环，开始填表中的信息
    for i in range(len(column_list)):
        column = column_list[i].split(' ')  # 将每个表头的信息分割
        tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=1).value = table_name  # 数据表的名称
        tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=2).value = column[0]  # 列名
        tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=3).value = column[1]  # 数据类型
        # 下面是数据的属性
        for key in column[2:]:
            if key == 'null':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=4).value = '1'  # 是否为空
            elif key == 'not_null':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=4).value = '0'  # 是否为空
            elif key == 'unique':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=5).value = '1'  # 是否为唯一
            elif key == 'pk':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=6).value = '1'  # 主键
            elif key == 'fk':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=7).value = '1'  # 外键
        column_names.append(column[0])  # 把所有列名存放起来
        for j in range(1, 8):  # 没有属性的信息赋值为NULL
            if tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=j).value is None:
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=j).value = 'Null'
    tbinfo.save('data/table_information.xlsx')  # 完成信息存放
    for i in range(length):  # 创建表头，挨个写入列名
        table.cell(row=1, column=i + 1).value = column_names[i]
    if table_name == 'authority':  # 当创建的表是authority表时，就给admin,root账户赋予权限
        table.cell(row=2, column=1).value = 'system'
        table.cell(row=3, column=1).value = 'table_information'
        max_row = table.max_row
        max_col = table.max_column
        for row in range(2, max_row + 1):
            for col in range(2, max_col + 1):
                table.cell(row=row, column=col).value = 'admin,root'
    current_db.save(db_path + current_dbname + '.xlsx')
    tbauth = load_workbook('data/table_authority.xlsx')  # 加载数据表的权限表
    tbauth_tb = tbauth[current_dbname]  # 读取当前数据表的信息表
    try:
        if user not in ['admin', 'root']:  # 授予权限
            auth_name = 'admin,root' + ',' + user  # 追加当前用户的权限，即此数据库的创建者
        else:
            auth_name = 'admin,root'
    except:
        auth_name = 'admin,root'
    tbauth_maxrow = tbauth_tb.max_row  # 读取最大列
    tbauth_maxcol = tbauth_tb.max_column  # 读取最大列
    tbauth_tb.cell(row=tbauth_maxrow + 1, column=1).value = current_dbname  # 写入数据库名称
    tbauth_tb.cell(row=tbauth_maxrow + 1, column=2).value = table_name  # 写入数据表名称
    for i in range(3, tbauth_maxcol + 1):
        tbauth_tb.cell(row=tbauth_maxrow + 1, column=i).value = auth_name  # 写入允许的用户名
        # authority_tbcol = ['database', 'table', 'select', 'insert', 'delete', 'update', 'use']  # 初始化权限项目
    if tbauth.worksheets[0].title == 'Sheet':  # 删除创建excel的默认sheet表
        del tbauth['Sheet']
    tbauth.save('data/table_authority.xlsx')  # 保存
    print('数据表 创建操作执行成功。')


def create_view(viewname, sql_view):
    '''
    创建视图
    CREATE VIEW IS_Student AS SELECT Sno,Sname,Sage FROM Student WHERE Sdept='IS
    :param viewname: 视图名称
    :param sql_view: sql语言
    :return:
    '''
    if not os.path.exists('data/view'):
        os.mkdir('data/view')
    if not os.path.exists('data/view/view_sql_information.xlsx'):  # 创立表的信息表
        Workbook().save('data/view/view_sql_information.xlsx')
    if not os.path.exists('data/view/' + using_dbname):
        os.mkdir('data/view/' + using_dbname)
    usedb = pd.read_excel('data/view/view_sql_information.xlsx', sheet_name=None)
    userdf = usedb[using_dbname]
    viewname_list = userdf['viewname'].tolist()
    if viewname in viewname_list:
        print('[!]此view名称已存在，请检查更换view名称！')
        return
    print('select start----------------------------------------------------------------')
    sss = sql_Analysis(sql_view, 'return')
    print('select end------------------------------------------------------------------')
    sss.to_excel('data/view/' + using_dbname + '/' + viewname + '.xlsx', sheet_name=viewname, index=False)
    try:
        db = load_workbook('data/view/view_sql_information.xlsx')
    except IOError:
        return
    if db.worksheets[0].title == 'Sheet':
        del db['Sheet']
    if using_dbname not in db.sheetnames:
        table = db.create_sheet(using_dbname)
        table.cell(row=1, column=1).value = 'viewname'
        table.cell(row=1, column=2).value = 'sql_view'
    try:
        table = db[using_dbname]
    except IOError:
        return
    max_row = table.max_row
    table.cell(row=max_row + 1, column=1).value = viewname
    table.cell(row=max_row + 1, column=2).value = sql_view
    db.save('data/view/view_sql_information.xlsx')
    print('视图' + viewname + '创建成功！')
    pass


def md5_pwd(pwd):
    """
    为了防止解密，hashlib.md5时加入了自己的字段
    将密码转为md5形式
    :param pwd: 密码铭文
    :return: 加密的结果
    """
    for i in range(pwd.shape[0]):
        values = pwd.iloc[i, 0]
        hash = hashlib.md5(bytes('dzx', encoding='utf8'))
        hash.update(bytes(values, encoding='utf8'))
        res = hash.hexdigest()
        pwd.iloc[i, 0] = res
    return pwd


def creat_index(sql, sql_word):
    # CREATE UNQUE INDEX SCno ON SC(Sno ASC,Cno DESC)
    Asql_word = [x.upper() for x in sql_word]  # 全变大写，方便查找
    try:  # 获取WHERE位置
        pos_index = Asql_word.index('INDEX')
    except:
        pos_index = -1
        print('[!]语法错误!：没有INDEX')
        return
    index_name = sql_word[pos_index + 1]
    table_name = re.findall('(.*)\(', sql_word[pos_index + 3])[0]
    columns = re.findall('\((.*)\)', sql)[0].split(',')
    columns = [x.split(' ') for x in columns]
    # print(index_name)
    # print(table_name)
    # print(columns)
    ind_df = pd.read_excel('data/' + using_dbname + '.xlsx', sheet_name=None)
    ind_tb = ind_df[table_name]
    reality_col = [x[0] for x in columns]
    # print(reality_col)
    asc_list = []  # 保存是升序还是降序的列表
    for lim in columns:  # 循环判断是否是升序降序
        if len(lim) == 1 or lim[1].upper() == 'ASC':  # 如果只有一个元素就是默认是升序，或者如果是ASC就是升序
            asc_list.append(1)  # 升序排序存入1
            continue
        if lim[1].upper() == 'DESC':  # 如果是DESC就是降序
            asc_list.append(0)  # 降序排序存入0
            continue
    # print(asc_list)
    sorted_ind_table = ind_tb.sort_values(by=reality_col,
                                          ascending=asc_list)  # 按照by_list(根据排序的列)和asc_list(排序顺序)排序
    # print(sorted_ind_table)
    ind_table = sorted_ind_table[[reality_col[0]]]
    for i in range(1, len(columns)):
        ind_test = sorted_ind_table[[reality_col[i]]]
        ind_table = pd.concat([ind_table, ind_test], axis=1)
    ind_table['ind_sum'] = ind_table[reality_col[0]]
    for i in range(1, len(reality_col)):
        ind_table['ind_sum'] = ind_table['ind_sum'].astype(str) + ind_table[reality_col[i]].astype(str)
    ind_table['hash_value'] = md5_pwd(ind_table[['ind_sum']])
    ind_table['pos_value'] = ind_table.index
    # print(ind_table)
    new_table = pd.DataFrame(ind_table['hash_value'])
    new_table = pd.concat([new_table, ind_table['pos_value']], axis=1)
    for col in reality_col:
        new_table = pd.concat([new_table, ind_table[col]], axis=1)
    # print(new_table)
    if not os.path.exists('data/index'):
        os.mkdir('data/index')
    if not os.path.exists('data/index/' + using_dbname):
        os.mkdir('data/index/' + using_dbname)
    new_table.to_excel('data/index/' + using_dbname + '/' + index_name + '.xlsx', index=False)
    print(table_name + '的索引' + index_name + '创建成功！')
    # CREATE UNQUE INDEX SCno ON SC(Sno ASC,Cno DESC)
    pass


def iter_cols(table):
    '''
    把表格中的数据按列提取
    :param table:
    :return: 列列表
    '''
    for row in table.iter_cols():
        yield [cell.value for cell in row]
    pass


def iter_rows(table):
    '''
    把表格中的数据按行提取
    :param table:
    :return: 行列表
    '''
    for col in table.iter_rows():
        yield [cell.value for cell in col]
    pass


def check_login(username, password):
    '''
    检查用户名与密码
    :param username: 用户名
    :param password: 密码
    :return:
    '''
    db = load_workbook('data/system.xlsx')  # 读取系统数据库里的用户表
    table = db['user']
    col_list = list(iter_cols(table))  # 按列读出数据
    try:
        pos = col_list[0].index(username)  # 读取用户名的位置
    except:
        return False
    right_pwd = col_list[1][pos]  # 根据用户名的位置读取相应的密码
    if hashlib.md5(password.encode('utf-8')).hexdigest() == right_pwd:  # 检验密码的md5值
        return True
    else:
        return False
    pass


def check_db_Authority(user, dbname, action):
    '''
    检查权限
    :param user:待检查的用户
    :param dbname: 用户需要访问的数据库
    :param action: 用户申请的操作
    :return:
    '''
    # global users_row, users_col
    db = load_workbook('data/system.xlsx')  # 读取系统数据库里面的权限表
    table = db['authority']
    max_row = table.max_row
    max_col = table.max_column
    # 查找对应的数据表和对应的操作的位置(对应的位置存放着允许执行这个表中的这个操作的用户名)
    for i in range(1, max_row + 1):
        if table.cell(row=i, column=1).value == dbname:
            users_row = i
            # print(users_row)
            break
    for j in range(1, max_col + 1):
        if table.cell(row=1, column=j).value == action:
            users_col = j
            # print(users_col)
            break
    # 读取允许执行这个表中的这个操作的用户名
    try:
        allow_users = table.cell(row=users_row, column=users_col).value.split(',')
    except:
        print('[!]没有在system/authority中找到' + dbname + '数据库')
        return False
    if user in allow_users:  # 如果列表中存在，则允许此操作
        print("您有权限对 {} 数据库进行 {} 操作!".format(dbname, action))
        return True
    else:  # 如果列表中不存在，则此操作不被允许
        print("您没有权限对 {} 数据库进行 {} 操作！请检查后重试。".format(dbname, action))
        return False
    pass


def check_table_Authority(user, dbname, tablename, operate):
    """
    检查权限
    :param user:待检查的用户
    :param dbname: 用户需要访问的数据库
    :param tablename: 用户需要访问的数据表
    :param operate: 用户申请的操作
    :return:
    """
    db = load_workbook('data/table_authority.xlsx')  # 读取系统数据库里面的权限表
    table = db[dbname]
    max_row = table.max_row
    max_col = table.max_column
    # 查找对应的数据表和对应的操作的位置(对应的位置存放着允许执行这个表中的这个操作的用户名)
    for i in range(1, max_row + 1):
        if table.cell(row=i, column=2).value == tablename:
            users_row = i
            # print(users_row)
            break
    for j in range(2, max_col + 1):
        if table.cell(row=1, column=j).value == operate:
            users_col = j
            # print(users_col)
            break
    # 读取允许执行这个表中的这个操作的用户名
    # print(users_row,users_col)
    try:
        allow_users = table.cell(row=users_row, column=users_col).value.split(',')
    except:
        print('[!]没有在table_authority中找到' + dbname + '数据库的' + tablename + '数据表')
        return False
    if user in allow_users:  # 如果列表中存在，则允许此操作
        print("您有 {} 权限操作 {} 数据库的 {} 数据表!".format(operate, dbname, tablename))
        return True
    else:  # 如果列表中不存在，则此操作不被允许
        print("[!]您没有 {} 权限操作 {} 数据库的 {} 数据表!请检查后重试。".format(operate, dbname, tablename))
        return False
    pass


def use_db(dbname):
    """
    使用数据库
    首先检查权限，
    然后加载此数据库
    :param dbname:
    :return:
    """
    global using_dbname
    global using_db
    if check_db_Authority(user, dbname, 'use'):
        using_dbname = dbname
        try:
            using_db = load_workbook(db_path + dbname + '.xlsx')
        except:
            print('[!]此数据库不存在，请检查后再次输入！')
            return
        print('数据库 ' + using_dbname + ' 已加载！')
    else:
        print('[!]权限不足！')
    pass


def unuse_db():
    """
    使用数据库
    首先检查权限，
    然后加载此数据库
    :param dbname:
    :return:
    """
    global using_dbname
    global using_db
    using_dbname = ''
    using_db = ''
    pass


def sql_Analysis(sql, tag=''):
    '''
    sql语法分析
    :param sql:
    :param tag:
    :return:
    '''
    sql = sql.strip()
    sql_word = sql.split(' ')  # 分割sql语句
    if len(sql_word) < 2:
        print('[!]语法错误')
        return
    operate = sql_word[0].lower()  # 操作名
    # print(sql_word)
    if operate == 'use':  # 操作名：use
        if sql_word[1] == 'database':  # 如果是use database
            use_db(sql_word[2])
            # try:
            #     use_db(sql_word[2])  # sql_word[2]是数据库名称
            # except:
            #     print('[!]打开数据库出错！')
        else:
            print("[!]语法错误！\neg:>use database dbname")
    elif operate == 'create':  # 操作名：create
        if check_db_Authority(user, using_dbname, operate) is False:
            return
        if sql_word[1].lower() == 'database':  # 如果是create database
            try:
                creat_db(sql_word[2])  # sql_word[2]是数据库名称
            except:
                print('[!]创建数据库出错！')
        elif sql_word[1].lower() == 'table':  # 如果是create table
            # print('验证表名：', sql_word[2])
            columns_list = re.findall('\((.*)\)', sql)[0].split(',')  # 按照正则表达式寻找()中语句
            print('新建 {} 数据库的列名为{}'.format(using_dbname, columns_list))
            try:
                creat_table(sql_word[2], using_db, using_dbname, columns_list)  # 创建数据表
            except:
                print('[!]创建数据表出错！')
        elif sql_word[1].lower() == 'view':  # 如果是create view
            # CREATE VIEW IS_Student AS SELECT Sno,Sname,Sage FROM Student WHERE Sdept='IS'
            viewname = sql_word[2]  # sql_word[2]是视图名称
            sql_view = ' '.join(sql_word[4:])  # 选择语句
            # print(sql_view)
            create_view(viewname, sql_view)
        elif sql_word[1].lower() == 'index' or sql_word[2].lower() == 'index':
            creat_index(sql, sql_word)
        else:
            print('[!]语法错误！请检查后重新输入。')
    elif operate == 'select':
        table_names = sql_word[3].split(',')
        if isinstance(table_names, list):
            for table_name in table_names:
                if check_table_Authority(user, using_dbname, table_name, operate) is False:
                    return
        else:
            print('[!]语法错误!：没有找到表名！')
            return
        if tag == 'return':
            return sl.sql_Analysis(using_dbname, sql, tag='return')
        else:
            sl.sql_Analysis(using_dbname, sql)
    elif operate == 'insert':
        table_name_info = sql_word[2]
        if '(' in table_name_info:  # 如果表名称后面有括号，就用正则表达式提取出数据表的名称和要插入列的名称
            table_name = re.findall('(.*)\(', table_name_info)[0]  # 提取表名
        else:
            table_name = table_name_info  # 如果没有括号，INTO数据就是表名
        if check_table_Authority(user, using_dbname, table_name, operate) is False:
            return
        ins.sql_Analysis(using_dbname, sql)
    elif operate == 'update':
        table_name = sql_word[1]
        if check_table_Authority(user, using_dbname, table_name, operate) is False:
            return
        ud.sql_Analysis(using_dbname, sql)
    elif operate == 'delete':
        table_name = sql_word[2]
        if check_table_Authority(user, using_dbname, table_name, operate) is False:
            return
        dl.sql_Analysis(using_dbname, sql)
    elif operate == 'grant':
        gt.sql_Analysis(using_dbname, sql)
    elif operate == 'revoke':
        rv.sql_Analysis(using_dbname, sql)
    elif operate == 'help':
        hl.sql_Analysis(using_dbname, sql)
    else:
        print('[!]语法错误！请检查后重新输入。')
    pass


def get_command():
    '''
    读取终端的命令
    :return:
    '''
    command = input("[{}@-]> ".format(user)) if not using_dbname else input("[{}@{}]> ".format(user, using_dbname))
    # command = input("[{}@-]> ".format('test root')) if not using_dbname else input(
    #     "[{}@{}]> ".format('test root', using_dbname))
    return command.strip()
    pass


def login():
    '''
    登录
    :return:
    '''
    global user  # 全局用户名
    print('--------- 请登录 ---------')
    # 输入用户名和密码
    username = input('username:')
    if username.lower() == 'logout':
        logout()
        return
    password = input('password:')
    # 检查信息
    if check_login(username, password):
        print('登录成功！欢迎 {}！'.format(username))
        user = username
    else:
        print('账号或密码错误，请尝试重新登录！')
        login()
        return


def logout():
    '''
    :return:
    '''
    main_choice()
    return


def register():
    usedb = pd.read_excel('data/system.xlsx', sheet_name=None)
    userdf = usedb['user']
    name_list = userdf['username'].tolist()
    sys_user = ['admin', 'root', 'logout']
    print('--------- 注册中 ---------')
    username = input('请输入您的 username:')
    if username.lower() == 'exit' or username.lower() == 'quit':
        main_choice()
        return
    elif username.lower() in sys_user:
        print('[!]您输入的用户名违规，请检查后重新注册！')
        register()
        return
    elif username in name_list:
        print('[!]您输入的用户名已存在，请检查后重新注册！')
        register()
        return
    password = input('请输入您的 password:')
    password_copy = input('请确认您的 password:')
    if password.lower() == 'exit' or password.lower() == 'quit':
        main_choice()
        return
    elif password == password_copy:
        usedb = load_workbook('data/system.xlsx')
        usertable = usedb['user']
        max_row = usertable.max_row
        usertable.cell(row=max_row + 1, column=1).value = username  # 写入账号
        usertable.cell(row=max_row + 1, column=2).value = hashlib.md5(
            password.encode('utf-8')).hexdigest()  # 写入密码并用MD5加密
        print('[*]账号注册成功！')
        print('您的用户名为：', username)
        print('您的密码为：', password)
        print('请妥善保存您的账号与密码！')
        usedb.save('data/system.xlsx')
        main_choice()
        return
    else:
        print('[!]您两次输入的密码不一致，请检查后重新注册！')
        register()
        return
    pass


def main_choice():
    print('Login or Exit or Register')
    print('1-登录\Login  2-退出\Exit  3-注册\Register')
    info = input('请输入您想操作的序号：')
    if info == '1':
        login()
        return
    elif info == '2':
        print('--------- 关机中 ---------')
        print('感谢您的使用！')
        exit(0)
    elif info == '3':
        register()
        return
    else:
        print('[!]输入错误，请检查后重新输入！')
        main_choice()
        return


def sql_Help():
    print("""
        ————————感谢您能使用我的DBMS！————————
    
       1.创建数据库：create database dbname
         create database S-T
       2.使用数据库：use database dbname
         use database S-T
       3.创建数据表：create table tbname (id int PK null,user char[10] )
         create table Student (Sno string,Sname string,Ssex string,Sage int,Sdept string)
       4.删除：DELETE FROM table_nmae WHERE column_name = 'Value'
         DELETE FROM Student WHERE Sage=20
       5.更新：UPDATE table_name SET column1=value1,column2=value2,... WHERE some_column=some_value;
         UPDATE Student SET Sage=22 WHERE Sno='201215121'
       6.插入：INSERT INTO table_name col1=val1,col2=val2&col3=val3,col4=val4
         INSERT INTO Student(Sno,Sname,Ssex,Sdept,Sage) VALUES('201215128','张成民','男','IS',18)
       7.查询：select a,b from table where c=x AND d=x（与）
              select a,b from table where c=x OR d=x （或）
              select a,b from table where c>x,d<x
              支持like，in，支持子查询
         SELECT Sno,Sname FROM Student WHERE Sdept='CS' AND Sage<20 group by Sno order by Sname ASC,Sno DESC
       8.权限：grant/revoke select on test_tb for testuser
       9.索引：creat view view_name as select xx from xx
       10.显示信息：help table/view/index
       """)
    pass


def Run():
    welcome()
    main_choice()
    while True:
        command = get_command()
        if command.lower() == 'exit' or command.lower() == 'quit':
            print('--------- 关机中 ---------')
            print('感谢您的使用！')
            exit(0)
        elif command.lower() == 'help':
            sql_Help()
        elif command.lower() == 'unuse database':
            unuse_db()
        elif command.lower() == 'logout':
            unuse_db()
            logout()
        else:
            sql_Analysis(command)


if __name__ == '__main__':
    Init()
    Run()
