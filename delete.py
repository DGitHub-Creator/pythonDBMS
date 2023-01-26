import os
import re

import pandas as pd
from openpyxl import *


def dl_equal(operator, lim, table):
    '''
    删除：条件：等于的
    :param operator: 操作符
    :param lim: 限制条件
    :param table: 现在要操作的数据表
    :return:
    '''
    limleft = lim[0]
    limright = re.split(r'[\']', lim[2])
    limright = [item for item in filter(lambda x: x != '', limright)][0]
    # print(type(limright))
    # print(limleft, operator, limright)
    try:
        limright = int(limright)
    except:
        try:
            limright = float(limright)
        except:
            limright = limright
    # print(table)
    try:
        # test = table[~((table[limleft]).astype(str).isin([limright]))]
        test = table[(table[limleft] != limright)]
    except:
        print('[!]语法错误!：没有找到', limleft, '列，请检查输入是否正确！')
        return
    return test


def dl_not_equal(operator, lim, table):
    '''
    删除：条件：不等于的
    :param operator: 操作符
    :param lim: 限制条件
    :param table: 现在要操作的数据表
    :return:
    '''
    limleft = lim[0]
    limright = re.split(r'[\']', lim[2])
    limright = [item for item in filter(lambda x: x != '', limright)][0]
    # limright = limright[0]
    # print(type(limright))
    # print(limleft, operator, limright)
    try:
        limright = int(limright)
    except:
        try:
            limright = float(limright)
        except:
            limright = limright
    try:
        # test = table[((table[limleft]).astype(str).isin([limright]))]
        test = table[(table[limleft] == limright)]
    except:
        print('[!]语法错误!：没有找到', limleft, '列，请检查输入是否正确！')
        return
    return test


def dl_less(operator, lim, table):
    '''
    删除：条件：小于的
    :param operator: 操作符
    :param lim: 限制条件
    :param table: 现在要操作的数据表
    :return:
    '''
    limleft = lim[0]
    limright = lim[2]
    # limright = int(limright)
    # print(limleft, operator, limright)
    try:
        limright = int(limright)
    except:
        try:
            limright = float(limright)
        except:
            limright = limright
    try:
        test = table[~(table[limleft] < limright)]
    except:
        print('[!]语法错误!：没有找到', limleft, '列，请检查输入是否正确！')
        return
    return test


def dl_more(operator, lim, table):
    '''
    删除：条件：大于的
    :param operator: 操作符
    :param lim: 限制条件
    :param table: 现在要操作的数据表
    :return:
    '''

    limleft = lim[0]
    limright = lim[2]
    # limright = int(limright)
    # print(limleft, operator, limright)
    try:
        limright = int(limright)
    except:
        try:
            limright = float(limright)
        except:
            limright = limright
    try:
        test = table[~(table[limleft] > limright)]
    except:
        print('[!]语法错误!：没有找到', limleft, '列，请检查输入是否正确！')
        return
    return test


def dl_less_equal(operator, lim, table):
    '''
    删除：条件：小于等于的
    :param operator: 操作符
    :param lim: 限制条件
    :param table: 现在要操作的数据表
    :return:
    '''

    limleft = lim[0]
    limright = lim[2]
    # limright = int(limright)
    # print(limleft, operator, limright)
    try:
        limright = int(limright)
    except:
        try:
            limright = float(limright)
        except:
            limright = limright
    try:
        test = table[~(table[limleft] <= limright)]
    except:
        print('[!]语法错误!：没有找到', limleft, '列，请检查输入是否正确！')
        return
    return test


def dl_more_equal(operator, lim, table):
    '''
    删除：条件：大于等于的
    :param operator: 操作符
    :param lim: 限制条件
    :param table: 现在要操作的数据表
    :return:
    '''
    limleft = lim[0]
    limright = lim[2]
    # limright = int(limright)
    # print(limleft, operator, limright)
    try:
        limright = int(limright)
    except:
        try:
            limright = float(limright)
        except:
            limright = limright
    try:
        test = table[~(table[limleft] >= limright)]
    except:
        print('[!]语法错误!：没有找到', limleft, '列，请检查输入是否正确！')
        return
    return test


def delete(using_dbname, table_name, con_limit):
    '''
    删除数据DELETE
    :param using_db: 数据库
    :param using_dbname: 数据库名称
    :param table_name: 数据表
    :param con_limit: 删除的限制条件
    :return:
    '''
    if using_dbname == '':  # 如果数据库名字为空，就报错
        print('[!]请选择一个数据库！')
        return
    else:
        using_db = pd.read_excel('data/' + using_dbname + '.xlsx', sheet_name=None)
        pass
    if table_name in using_db.keys():
        try:
            table = using_db[table_name]  # 获取数据表
        except:
            print('[!]查无此表！', table_name, '表没有找到')
            return
    elif os.path.exists('data/view/' + using_dbname + '/' + table_name + '.xlsx'):
        try:
            view_db = pd.read_excel('data/view/' + using_dbname + '/' + table_name + '.xlsx', sheet_name=None)  # 获取数据表
            table = view_db[table_name]
        except:
            print('[!]查无此表！', table_name, '表没有找到')
            return
    else:
        print('[!]查无此表！', table_name, '表没有找到')
        return
    # print(table)
    if len(con_limit) == 0:
        bookdb = load_workbook('data/' + using_dbname + '.xlsx')
        booktable = bookdb[table_name]
        book_maxrow = booktable.max_row
        booktable.delete_rows(2, book_maxrow)
        bookdb.save('data/' + using_dbname + '.xlsx')
        print("[*]删除成功！删除了", book_maxrow - 1, "个元组")
        return
    else:
        predicate = con_limit[-1]
        # predicate = ' ',AND,OR,IN,NOTIN,BETWEENAND,NOTBETWEENAND.LIKE,NOTLIKE
        # print('predicate:', predicate)
        old_table = table
        new_table = table
        # DELETE FROM Student WHERE Sage=20
        # DELETE FROM Student WHERE Ssex='男'
        if predicate == ' ':
            limit = con_limit[:-1]
            # print(limit)
            # 删除 =条件
            for lim in limit:
                operator = lim[1]
                # DELETE FROM Student WHERE Sage=18
                if operator == '=':
                    new_table = dl_equal(operator, lim, table)
                # DELETE FROM Student WHERE Sage<18
                if operator == '<':
                    new_table = dl_less(operator, lim, table)
                # DELETE FROM Student WHERE Sage>18
                if operator == '>':
                    new_table = dl_more(operator, lim, table)
                # DELETE FROM Student WHERE Sage<=18
                if operator == '<=':
                    new_table = dl_less_equal(operator, lim, table)
                # DELETE FROM Student WHERE Sage>=18
                if operator == '>=':
                    new_table = dl_more_equal(operator, lim, table)
                # DELETE FROM Student WHERE Sage<>18
                if operator == '!=' or operator == '<>':
                    new_table = dl_not_equal(operator, lim, table)
        # DELETE FROM Student WHERE Sage=20 OR Ssex='男'
        if predicate == 'OR':
            limit = con_limit[:-1]
            # print(limit)
            # 删除 =条件
            for lim in limit:
                operator = lim[1]
                # DELETE FROM Student WHERE Sage=18
                if operator == '=':
                    new_table = dl_equal(operator, lim, new_table)
                # DELETE FROM Student WHERE Sage<18
                if operator == '<':
                    new_table = dl_less(operator, lim, new_table)
                # DELETE FROM Student WHERE Sage>18
                if operator == '>':
                    new_table = dl_more(operator, lim, new_table)
                # DELETE FROM Student WHERE Sage<=18
                if operator == '<=':
                    new_table = dl_less_equal(operator, lim, new_table)
                # DELETE FROM Student WHERE Sage>=18
                if operator == '>=':
                    new_table = dl_more_equal(operator, lim, new_table)
                # DELETE FROM Student WHERE Sage<>18
                if operator == '!=' or operator == '<>':
                    new_table = dl_not_equal(operator, lim, new_table)
        # DELETE FROM Student WHERE Sage=20 AND Ssex='男'
        # DELETE FROM Student WHERE Sage<=20 AND Ssex='女'
        if predicate == 'AND':
            limit = con_limit[:-1]
            # print(limit)
            # 删除 =条件
            table_copy = table
            for lim in limit:
                operator = lim[1]
                # DELETE FROM Student WHERE Sage=18
                if operator == '=':
                    table_copy = dl_not_equal(operator, lim, table_copy)
                # DELETE FROM Student WHERE Sage<18
                if operator == '<':
                    table_copy = dl_more_equal(operator, lim, table_copy)
                # DELETE FROM Student WHERE Sage>18
                if operator == '>':
                    table_copy = dl_less_equal(operator, lim, table_copy)
                # DELETE FROM Student WHERE Sage<=18
                if operator == '<=':
                    table_copy = dl_more(operator, lim, table_copy)
                # DELETE FROM Student WHERE Sage>=18
                if operator == '>=':
                    table_copy = dl_less(operator, lim, table_copy)
                # DELETE FROM Student WHERE Sage<>18
                if operator == '!=' or operator == '<>':
                    table_copy = dl_equal(operator, lim, table_copy)
            if table_copy is None:
                return
            # print(list(table_copy.index))
            new_table = table.drop(index=list(table_copy.index))
            # print(new_table)
        # DELETE FROM Student WHERE Sdept in ('CS','MA')
        if predicate == 'IN':
            column = con_limit[0]
            # print(column, end=' ')
            limit = con_limit[1]
            # print(limit)
            try:
                new_table = table[~table[column].isin(limit)]
            except:
                print('[!]语法错误!：没有找到', column, '列，请检查输入是否正确！')
                return
            # print(new_table)
        # DELETE FROM Student WHERE Sdept not in ('CS','MA')
        if predicate == 'NOTIN':
            column = con_limit[0]
            # print(column, end=' ')
            limit = con_limit[1]
            # print(limit)
            try:
                new_table = table[table[column].isin(limit)]
            except:
                print('[!]语法错误!：没有找到', column, '列，请检查输入是否正确！')
                return
            # print(new_table)
        # DELETE FROM Student WHERE Sage between 18 and 19
        if predicate == 'BETWEENAND':
            column = con_limit[0]
            minn = int(con_limit[1][0])
            maxx = int(con_limit[1][1])
            limit = list(range(minn, maxx + 1))
            # print(limit)
            try:
                new_table = table[~table[column].isin(limit)]
            except:
                print('[!]语法错误!：没有找到', column, '列，请检查输入是否正确！')
                return
            # print(new_table)
        # DELETE FROM Student WHERE Sage not between 18 and 19
        if predicate == 'NOTBETWEENAND':
            column = con_limit[0]
            minn = int(con_limit[1][0])
            maxx = int(con_limit[1][1])
            limit = list(range(minn, maxx + 1))
            # print(limit)
            try:
                new_table = table[table[column].isin(limit)]
            except:
                print('[!]语法错误!：没有找到', column, '列，请检查输入是否正确！')
                return
        # DELETE FROM Student WHERE Sname LIKE '刘%'
        if predicate == 'LIKE':  # 如果谓词是LIKE
            column = con_limit[0]
            limit = con_limit[1]
            if con_limit[2] != 'LIKE':  # 不是LIKE就是通配符
                escape = con_limit[2][-2]
                wc_str = limit[0].replace('%', '(.*)').replace('_', '(.)').replace(escape + '(.*)', '%').replace(
                    escape + '(.)', '_')
            else:
                wc_str = limit[0].replace('%', '(.*)').replace('_', '(.)')
            try:
                col_list = table[column].tolist()
            except:
                print('[!]语法错误!：没有找到', column, '列')
                return
            name_list = []
            for i in col_list:
                s = re.search(wc_str, i)
                try:
                    name_list.append(s[0])
                except:
                    ''
            try:
                new_table = table[~table[column].isin(name_list)]
            except:
                print('[!]语法错误!：没有找到', column, '列，请检查输入是否正确！')
                return
            # print(new_table)
        # print(new_table)
        # 保存
        print("[*]删除成功！删除了", old_table.shape[0] - new_table.shape[0], "个元组")

        book = load_workbook('data/' + using_dbname + '.xlsx')
        writer = pd.ExcelWriter('data/' + using_dbname + '.xlsx')
        writer.book = book
        # 保存操作 保存修改后的表
        new_table.to_excel(writer, sheet_name=table_name, index=False)
        writer.save()
        # 因为pandas没有办法覆盖数据表，它会新建一个数据表，并在你所设置的名字后面加一个1，
        # 这样此数据表的名称变为:table_name1,所以要重新读入，删除旧的表并且把新的表改名
        wb = load_workbook('data/' + using_dbname + '.xlsx')  # 读入数据库
        del wb[table_name]  # 删除掉旧的表
        for sheet in wb:  # 循环查找table_name1，改成table_name
            if sheet.title == table_name + '1':
                sheet.title = table_name
                break
        wb.save('data/' + using_dbname + '.xlsx')  # 把修改后的数据表保存

        print(table_name, '表删除操作完成！')
        return


def sql_Analysis(using_dbname, sql, tag=''):
    sql = sql.strip()
    sql_word = sql.split(' ')  # 分割sql语句
    if len(sql_word) < 2:
        print('[!]语法错误')
        return
    operate = sql_word[0].lower()  # 操作名
    # DELETE FROM SC WHERE Sdept='CS' AND Sage<20
    # group by sno order by sname
    if operate == 'delete':
        pos_where = -1
        # print(operate.upper())
        # columns = sql_word[1]
        # print(columns)
        if sql_word[1].lower() != 'from':
            print('[!]语法错误!：FROM使用错误！请检查FROM是否存在！')
            return
        # print(sql_word[1], end=' ')
        table_name = sql_word[2]
        # print(table_name)
        Asql_word = [x.upper() for x in sql_word]
        try:
            pos_where = Asql_word.index('WHERE')
        except:
            pos_where = -1
        if pos_where == -1:
            if len(sql_word) > 3:
                print('[!]语法错误!：WHERE使用错误！请检查WHERE是否存在！')
                return
        # where>group>order
        con_limit = []
        limit_where = []
        if pos_where != -1:
            pos_end = len(sql_word)
            # print(sql_word[3], end=' ')
            limit_where = sql_word[pos_where + 1:pos_end]
            # print(limit_where)
            subqueries = 'AND'
            # DELETE FROM SC WHERE Sdept='CS' AND Sage<=20
            con_limit = []  # 分割限制条件
            # AND
            if 'BETWEEN' not in Asql_word and 'AND' in Asql_word:
                subqueries = 'AND'
                for i in range(0, len(limit_where), 2):
                    con_limit.append(re.split(r'(>=|<=|!=|<>|!>|!<|=|<|>)', limit_where[i]))
                    # print(con_limit)
                con_limit.append('AND')
            # print(con_limit)
            # OR
            # DELETE FROM SC WHERE Sdept='CS' or Sage<=20
            if 'OR' in Asql_word:
                subqueries = 'OR'
                for i in range(0, len(limit_where), 2):
                    con_limit.append(re.split(r'(>=|<=|!=|<>|!>|!<|=|<|>)', limit_where[i]))
                    # print(con_limit)
                con_limit.append('OR')
            # print(con_limit)
            # IN
            # DELETE FROM SC WHERE Sdept in ('CS','MA','IS')
            if 'NOT' not in Asql_word and 'IN' in Asql_word:
                subqueries = 'IN'
                con_1 = limit_where[0]
                con_2 = re.split(r'[(|)|\'|,]', limit_where[2])
                con_2 = [item for item in filter(lambda x: x != '', con_2)]
                con_limit.append(con_1)
                con_limit.append(con_2)
                con_limit.append('IN')
            # print(con_limit)
            # NOT IN
            # DELETE FROM SC WHERE Sdept not in ('CS','MA','IS')
            if 'NOT' in Asql_word and 'IN' in Asql_word:
                subqueries = 'NOTIN'
                con_1 = limit_where[0]
                con_2 = re.split(r'[(|)|\'|,]', limit_where[3])
                con_2 = [item for item in filter(lambda x: x != '', con_2)]
                con_limit.append(con_1)
                con_limit.append(con_2)
                con_limit.append('NOTIN')
            # print(con_limit)
            # BETWEEN AND
            # DELETE FROM SC WHERE Sage between 20 and 23
            if 'NOT' not in Asql_word and 'BETWEEN' in Asql_word and 'AND' in Asql_word:
                subqueries = 'BETWEENAND'
                con_1 = limit_where[0]

                if int(limit_where[2]) <= int(limit_where[4]):
                    con_2 = [limit_where[2], limit_where[4]]
                else:
                    print('BETWEEN AND的MIN,MAX值错误')
                    return
                con_limit.append(con_1)
                con_limit.append(con_2)
                con_limit.append('BETWEENAND')
            # print(con_limit)
            # NOT BETWEEN AND
            # DELETE FROM SC WHERE Sage not between 20 and 23
            if 'NOT' in Asql_word and 'BETWEEN' in Asql_word and 'AND' in Asql_word:
                subqueries = 'NOTBETWEENAND'
                con_1 = limit_where[0]
                # con_2 = [limit_where[3], limit_where[5]]
                if int(limit_where[3]) <= int(limit_where[5]):
                    con_2 = [limit_where[3], limit_where[5]]
                else:
                    print('BETWEEN AND的MIN,MAX值错误')
                    return
                con_limit.append(con_1)
                con_limit.append(con_2)
                con_limit.append('NOTBETWEENAND')
            # print(con_limit)
            # LIKE
            # DELETE FROM SC WHERE Sname LIKE '刘%'
            if 'NOT' not in Asql_word and 'LIKE' in Asql_word:
                subqueries = 'LIKE'
                con_1 = limit_where[0]
                con_2 = re.split(r'[(|)|\'|,]', limit_where[2])
                con_2 = [item for item in filter(lambda x: x != '', con_2)]
                con_limit.append(con_1)
                con_limit.append(con_2)
                con_limit.append('LIKE')
            # print(con_limit)
            # NOT LIKE
            # DELETE FROM SC WHERE Sname NOT LIKE '刘%'
            if 'NOT' in Asql_word and 'LIKE' in Asql_word:
                subqueries = 'NOTLIKE'
                con_1 = limit_where[0]
                con_2 = re.split(r'[(|)|\'|,]', limit_where[3])
                con_2 = [item for item in filter(lambda x: x != '', con_2)]
                con_limit.append(con_1)
                con_limit.append(con_2)
                con_limit.append('NOTLIKE')
            # print(con_limit)
            # IS NULL
            # DELETE FROM SC WHERE Sname IS NULL
            if 'IS' in Asql_word and 'NOT' not in Asql_word and 'NULL' in Asql_word:
                subqueries = 'ISNULL'
                con_1 = limit_where[0]
                con_limit.append(con_1)
                con_limit.append('ISNULL')
            # print(con_limit)
            # IS NOT NULL
            # DELETE FROM SC WHERE Sname IS NOT NULL
            if 'IS' in Asql_word and 'NOT' in Asql_word and 'NULL' in Asql_word:
                subqueries = 'ISNOTNULL'
                con_1 = limit_where[0]
                con_limit.append(con_1)
                con_limit.append('ISNOTNULL')
            if len(con_limit) == 0:
                # con_limit = limit_where
                for i in range(0, len(limit_where), 2):
                    con_limit.append(re.split(r'(>=|<=|!=|<>|!>|!<|=|<|>)', limit_where[i]))
                con_limit.append(' ')
            # print(con_limit)
        delete(using_dbname, table_name, con_limit)


if __name__ == '__main__':
    for i in range(10):
        using_dbname = 'S-T'
        # using_db = pd.read_excel('data/' + using_dbname + '.xlsx', sheet_name=None)
        sql = input('[!][' + str(i) + ']>> ')
        sql_Analysis(using_dbname, sql)
